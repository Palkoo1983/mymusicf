import os
from openpyxl import Workbook, load_workbook
from datetime import datetime

def log_order(data: dict, excel_dir='admin_logs'):
    today = datetime.now().strftime('%Y-%m-%d')
    filename = f"rendelesek_{today}.xlsx"
    filepath = os.path.join(excel_dir, filename)
    os.makedirs(excel_dir, exist_ok=True)

    headers = [
        'email', 'style', 'vocal', 'language',
        'event_type', 'brief', 'lyrics',
        'mp3_link_1', 'mp3_link_2',
        'date', 'time'
    ]

    now = datetime.now()
    row = [
        data.get('email', ''),
        data.get('style', ''),
        data.get('vocal', ''),
        data.get('language', ''),
        data.get('event_type', ''),
        data.get('brief', ''),
        data.get('lyrics', ''),
        data.get('mp3_link_1', ''),
        data.get('mp3_link_2', ''),
        now.strftime('%Y-%m-%d'),
        now.strftime('%H:%M:%S')
    ]

    if os.path.exists(filepath):
        wb = load_workbook(filepath)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.append(headers)

    ws.append(row)
    wb.save(filepath)